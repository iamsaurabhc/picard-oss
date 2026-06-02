import json
import uuid
import pytest

from app.db.models import Chunk, Document, Entity, EntityMention, MetadataTag, TabularCell, TabularReview, Workspace
from app.db.session import utc_now_iso
from app.services.fts_search import FtsHit
from app.services.tabular_extractor import TabularCellExtraction, _parse_llm_json, extract_cell
from app.schemas import TabularColumn


@pytest.fixture()
def tabular_fixture(db_session):
    ws_id = str(uuid.uuid4())
    now = utc_now_iso()
    db_session.add(Workspace(id=ws_id, name="TR Test", matter_ref=None, created_at=now, updated_at=now))
    db_session.flush()
    doc_id = str(uuid.uuid4())
    db_session.add(
        Document(
            id=doc_id,
            workspace_id=ws_id,
            file_name="contract.pdf",
            local_path="/tmp/contract.pdf",
            parse_status="done",
            created_at=now,
        )
    )
    db_session.flush()
    chunk_id = str(uuid.uuid4())
    db_session.add(
        Chunk(
            id=chunk_id,
            document_id=doc_id,
            page_number=2,
            chunk_type="paragraph",
            bbox_json='{"x":0.1,"y":0.2,"w":0.5,"h":0.1}',
            text_content="The governing law shall be the laws of New York.",
            heading_path=None,
            section_key=None,
            token_count=10,
        )
    )
    db_session.flush()
    review_id = str(uuid.uuid4())
    columns = [
        TabularColumn(
            key="governing_law",
            label="Governing Law",
            format="text",
            prompt="State the governing law.",
        )
    ]
    db_session.add(
        TabularReview(
            id=review_id,
            workspace_id=ws_id,
            title="Pilot",
            columns_config_json=json.dumps([c.model_dump() for c in columns]),
            document_ids_json=json.dumps([doc_id]),
            created_at=now,
        )
    )
    cell_id = str(uuid.uuid4())
    db_session.add(
        TabularCell(
            id=cell_id,
            review_id=review_id,
            document_id=doc_id,
            column_key="governing_law",
            status="pending",
        )
    )
    db_session.commit()
    return {
        "ws_id": ws_id,
        "doc_id": doc_id,
        "chunk_id": chunk_id,
        "review_id": review_id,
        "cell_id": cell_id,
        "columns": columns,
    }


def test_create_and_get_review(client, db_session, tabular_fixture):
    f = tabular_fixture
    r = client.post(
        "/tabular/reviews",
        json={
            "workspace_id": f["ws_id"],
            "title": "Second Review",
            "columns": [{"key": "term", "label": "Term", "format": "text", "prompt": "State the term."}],
            "document_ids": [f["doc_id"]],
        },
    )
    assert r.status_code == 201
    data = r.json()
    assert data["title"] == "Second Review"
    assert len(data["cells"]) == 1

    r2 = client.get(f"/workspaces/{f['ws_id']}/tabular/reviews")
    assert r2.status_code == 200
    assert len(r2.json()) >= 2


def test_extract_cell_mock_llm(client, db_session, tabular_fixture, monkeypatch):
    f = tabular_fixture
    review = db_session.get(TabularReview, f["review_id"])
    cell = db_session.get(TabularCell, f["cell_id"])
    column = f["columns"][0]

    payload = TabularCellExtraction(
        summary="New York Law",
        reasoning="Chunk states New York governing law.",
        chunk_ids=[f["chunk_id"]],
        flag="green",
    ).model_dump_json()

    monkeypatch.setattr("app.services.tabular_extractor.completion", lambda **kwargs: payload)

    def _fake_hits(db, **kwargs):
        return [
            FtsHit(
                chunk_id=f["chunk_id"],
                document_id=f["doc_id"],
                page_number=2,
                text_content="The governing law shall be the laws of New York.",
                heading_path=None,
                section_key=None,
                bbox_json="{}",
                score=-1.0,
            )
        ]

    monkeypatch.setattr("app.services.tabular_extractor.gather_cell_chunks", _fake_hits)

    out = extract_cell(db_session, cell=cell, review=review, column=column)
    assert out.status == "done"
    assert "New York" in (out.summary or "")
    assert f["chunk_id"] in out.source_chunk_ids
    assert "[[page:2||quote:" in (out.summary or "")


def test_regenerate_single_cell(client, tabular_fixture, monkeypatch):
    f = tabular_fixture
    payload = TabularCellExtraction(
        summary="Updated",
        reasoning="ok",
        chunk_ids=[f["chunk_id"]],
        flag="green",
    ).model_dump_json()
    monkeypatch.setattr("app.services.tabular_extractor.completion", lambda **kwargs: payload)
    monkeypatch.setattr(
        "app.services.tabular_extractor.gather_cell_chunks",
        lambda db, **kwargs: [
            FtsHit(
                chunk_id=f["chunk_id"],
                document_id=f["doc_id"],
                page_number=2,
                text_content="text",
                heading_path=None,
                section_key=None,
                bbox_json="{}",
                score=-1.0,
            )
        ],
    )

    r = client.post(f"/tabular/cells/{f['cell_id']}/regenerate")
    assert r.status_code == 200
    assert r.json()["summary"] is not None


def test_parties_extraction_uses_early_page_chunk(db_session, monkeypatch):
    ws_id = str(uuid.uuid4())
    now = utc_now_iso()
    db_session.add(Workspace(id=ws_id, name="Parties", matter_ref=None, created_at=now, updated_at=now))
    db_session.flush()
    doc_id = str(uuid.uuid4())
    db_session.add(
        Document(
            id=doc_id,
            workspace_id=ws_id,
            file_name="cci.pdf",
            local_path="/tmp/cci.pdf",
            parse_status="done",
            created_at=now,
        )
    )
    db_session.flush()
    early_chunk = str(uuid.uuid4())
    db_session.add(
        Chunk(
            id=early_chunk,
            document_id=doc_id,
            page_number=1,
            chunk_type="paragraph",
            bbox_json="{}",
            text_content="Google LLC and Informant-1 are opposite parties in Case No. 06 of 2014.",
            heading_path=None,
            section_key=None,
            token_count=15,
        )
    )
    entity_id = str(uuid.uuid4())
    db_session.add(
        Entity(
            id=entity_id,
            workspace_id=ws_id,
            entity_type="party",
            canonical_value="google llc",
            display_value="Google LLC",
        )
    )
    db_session.add(
        EntityMention(
            id=str(uuid.uuid4()),
            entity_id=entity_id,
            document_id=doc_id,
            chunk_id=early_chunk,
            page_number=1,
            char_start=None,
            char_end=None,
            surface_text="Google LLC",
            confidence=0.9,
            source="rule",
        )
    )
    db_session.add(
        MetadataTag(
            id=str(uuid.uuid4()),
            document_id=doc_id,
            tag_key="party_1",
            tag_value="Google LLC",
            source_chunk_id=early_chunk,
        )
    )
    review_id = str(uuid.uuid4())
    column = TabularColumn(
        key="parties",
        label="Parties",
        format="bulleted_list",
        prompt="List all parties.",
    )
    db_session.add(
        TabularReview(
            id=review_id,
            workspace_id=ws_id,
            title="Parties review",
            columns_config_json=json.dumps([column.model_dump()]),
            document_ids_json=json.dumps([doc_id]),
            created_at=now,
        )
    )
    cell_id = str(uuid.uuid4())
    db_session.add(
        TabularCell(
            id=cell_id,
            review_id=review_id,
            document_id=doc_id,
            column_key="parties",
            status="pending",
        )
    )
    db_session.commit()

    payload = TabularCellExtraction(
        summary="• Google LLC — opposite party\n• Informant-1 — informant",
        reasoning="From page 1 caption.",
        chunk_ids=[early_chunk],
        flag="green",
    ).model_dump_json()
    monkeypatch.setattr("app.services.tabular_extractor.completion", lambda **kwargs: payload)

    review = db_session.get(TabularReview, review_id)
    cell = db_session.get(TabularCell, cell_id)
    out = extract_cell(db_session, cell=cell, review=review, column=column)
    assert out.status == "done"
    assert early_chunk in out.source_chunk_ids
    assert "Google" in (out.summary or "")


def test_export_strips_citations(client, db_session, tabular_fixture):
    f = tabular_fixture
    cell = db_session.get(TabularCell, f["cell_id"])
    cell.status = "done"
    cell.summary = 'New York Law [[page:2||quote:sample text]]'
    db_session.commit()

    r = client.get(f"/tabular/reviews/{f['review_id']}/export.xlsx")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    values = [cell.value for row in ws.iter_rows(min_row=2, max_row=2) for cell in row]
    joined = " ".join(str(v) for v in values if v)
    assert "[[" not in joined
    assert "New York" in joined


def test_parse_llm_json_coerces_list_summary_to_bullets():
    raw = json.dumps(
        {
            "summary": [
                "Google LLC: Opposite Party",
                "Informant-1: Informant",
            ],
            "reasoning": "From caption.",
            "chunk_ids": ["abc-123"],
            "flag": "green",
        }
    )
    parsed = _parse_llm_json(raw)
    assert parsed is not None
    assert "Google LLC" in parsed.summary
    assert parsed.summary.startswith("- ")
    assert "\n" in parsed.summary
