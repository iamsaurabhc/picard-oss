from __future__ import annotations

import io
import json
import re

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import Document, TabularCell, TabularReview
from app.schemas import TabularColumn

_CITATION_MARKER_RE = re.compile(r"\[\[[^\]]+\]\]")
_NUMERIC_CITATION_RE = re.compile(r"\[\d+\]")


def _strip_citations(text: str | None) -> str:
    if not text:
        return ""
    out = _CITATION_MARKER_RE.sub("", text)
    out = _NUMERIC_CITATION_RE.sub("", out)
    return re.sub(r"[ \t]+", " ", out).strip()


def _parse_columns(raw: str) -> list[TabularColumn]:
    data = json.loads(raw)
    return [TabularColumn(**item) for item in data]


def build_review_xlsx(db: Session, review_id: str) -> bytes:
    review = db.get(TabularReview, review_id)
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")

    columns = _parse_columns(review.columns_config_json)
    document_ids: list[str] = json.loads(review.document_ids_json)
    documents = {
        d.id: d
        for d in db.scalars(select(Document).where(Document.id.in_(document_ids))).all()
    }
    cells = db.scalars(select(TabularCell).where(TabularCell.review_id == review_id)).all()
    cell_map = {(c.document_id, c.column_key): c for c in cells}

    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    headers = ["Document"] + [c.label for c in columns]
    ws.append(headers)
    wrap = Alignment(wrap_text=True, vertical="top")

    for doc_id in document_ids:
        doc = documents.get(doc_id)
        if not doc:
            continue
        row = [doc.file_name]
        for col in columns:
            cell = cell_map.get((doc_id, col.key))
            if not cell or cell.status in ("pending", "generating"):
                row.append("")
            elif cell.status == "error":
                row.append("Error")
            else:
                row.append(_strip_citations(cell.summary))
        ws.append(row)
        row_idx = ws.max_row
        for col_idx in range(2, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = wrap

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
